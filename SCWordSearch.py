import os
import re
from openpyxl import Workbook

directions = ["go","leave", "move", "abscond", "depart", "hightail", "journey","split",
"set","ready", "roll",
"driving","drive", "cruise", "cruising",
"road","avenue", "boulevard", "course", "highway", "lane",
"miles", "kilometers","mile","kilometer","distance",
"left", "right", "continue", "straight","turn", "roundabout", "ahead",
"first", "second", "third", "uno", "dos", "tres", 
"police", "detective",
"crash", "accident","issue", "calamity", "collision", "hazard",
 "traffic", "slow down", "gridlock",
 "speed", "pace", "velocity", "quickness", 
 "and then", "after", "afterwards"
 "rerouting", "route", "arrived", "there", "arrive", "made it"]

wb = Workbook()

testDirections = ["go", "crash"]
# for file in os.listdir("/Users/wadeglaser/Desktop/SCS"):
#  	if file.endswith(".txt"):
#  		with open(file,'r') as f:
#  			content = f.read()
#  			#i = content.find("leave")
#  			#if(i>0):
#  			#	print("FOUND IN: " + file)
#  			#	print(content[i-20:i+20])
#  			for word in testDirections:
#  				ws = wb.create_sheet(word)
#  				arr= [m.start() for m in re.finditer(word, content)]
#  				print("FILE: "+ file+"*****************************")
#  				for x in arr:
#  					print(content[x-20:x+20])

for word in directions:
	ws = wb.create_sheet(word)

	for file in os.listdir("/Users/wadeglaser/Desktop/SCS"):
		if file.endswith(".txt"):
			with open(file,'r') as f:
  				content = f.read()	
  				arr= [m.start() for m in re.finditer(word, content)]
  				r = 1;
  				for sentence in arr:
  					ws.cell(row=r, column = 1, value = content[sentence-40:sentence+40])
  					ws.cell(row=r,column=2, value = file)
  					r=r+1





wb.save('words.xlsx')
